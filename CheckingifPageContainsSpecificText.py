import requests
from bs4 import BeautifulSoup


#r = requests.get('http://www.fveconstruction.ch/anDetails.asp?RT=2&M=01&R=1&ID=42105701')
#
#if 'not' in r.text:
#    print ('Yes')
#else:
#    print ('No')

# import module
import openpyxl
  
# load excel with its path
wrkbk = openpyxl.load_workbook("Book1.xlsx")
  
sh = wrkbk.active
      
    # iterate through excel and display data
for i in range(1, sh.max_row+1):
    print("\n")
    print("Row ", i, " data :")
      
    for j in range(1, sh.max_column+1):
        cell_obj = sh.cell(row=i, column=j)
        print(cell_obj.value, end=" ")